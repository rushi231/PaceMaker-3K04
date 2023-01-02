from openpyxl import Workbook #pip install openpyxl
from openpyxl import load_workbook

wb = load_workbook("AccountData.xlsx")

ws = wb["mun"]
AOOLRL = ws["B4"]
print(AOOLRL.value)


